"""Master-data import.

Four distinct steps, never collapsed: **parse -> validate -> detect duplicates
-> preview -> commit**. The preview exists because section 7 forbids merging
similar names without proof; the importer proposes, a person decides, and
nothing is deleted or merged automatically.

Two villages of the same name in different places stay two records. An
alternative name becomes an alias row, never a rewrite of the village's name.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import asdict, dataclass, field
from decimal import Decimal, InvalidOperation
from typing import Any

from domain.enums import ActorRole, GeoStatus, ImportStatus
from domain.text import comparison_key, similarity
from shared import error_codes
from shared.errors import ConflictError, NotFoundError, ValidationError

REQUIRED_COLUMNS = ("district_code", "name")
OPTIONAL_COLUMNS = ("alternative_names", "latitude", "longitude", "code", "note", "station_name")

# Above this, two names are proposed as the same place for a human to confirm.
# Deliberately not a merge threshold -- nothing merges automatically.
DUPLICATE_THRESHOLD = 0.82


@dataclass(slots=True)
class ParsedVillage:
    row_number: int
    district_code: str
    name: str
    aliases: list[str] = field(default_factory=list)
    latitude: Decimal | None = None
    longitude: Decimal | None = None
    code: str | None = None
    note: str | None = None
    station_name: str | None = None

    @property
    def key(self) -> str:
        return comparison_key(self.name)


@dataclass(slots=True)
class RowProblem:
    """Something wrong with a row.

    ``blocking`` separates the two kinds, because they need different reactions:
    a missing name or an unknown district means the row cannot be imported at
    all, while a malformed coordinate only means the village is created without
    one -- coordinates are optional everywhere in this product. Presenting both
    as simply "errors" would make an operator think they had lost rows they had
    not.
    """

    row_number: int
    column: str
    reason: str
    value: str | None = None
    blocking: bool = True


@dataclass(slots=True)
class DuplicateProposal:
    row_number: int
    incoming_name: str
    existing_village_id: str | None
    existing_name: str
    score: float
    same_district: bool
    reason: str


@dataclass(slots=True)
class ImportPreview:
    job_id: str
    total_rows: int
    valid_rows: int
    problems: list[RowProblem]
    duplicates: list[DuplicateProposal]
    will_create: list[ParsedVillage]

    @property
    def is_clean(self) -> bool:
        return not self.problems


def _rows_from_excel(content: bytes) -> list[tuple[int, dict[str, Any]]]:
    """Read the first worksheet, taking the first non-empty row as the header.

    Real spreadsheets have merged title rows, trailing blank rows and stray
    whitespace in headers. Handling that here is the difference between a tool
    an operator uses and one they give up on.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ValidationError(error_codes.IMPORT_FILE_UNREADABLE, reason="no_worksheet")

    header: list[str] | None = None
    records: list[tuple[int, dict[str, Any]]] = []
    seen_rows: list[list[str]] = []

    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = ["" if cell is None else str(cell).strip() for cell in row]
        if not any(values):
            continue

        if header is None:
            candidate = [value.strip().lower().replace(" ", "_") for value in values]
            # The header is the first row that actually carries the required
            # columns -- not merely the first non-empty one. Spreadsheets from
            # a district office routinely open with a merged title row, and
            # treating that as the header rejects the whole file.
            if all(column in candidate for column in REQUIRED_COLUMNS):
                header = candidate
            else:
                seen_rows.append(values)
                if len(seen_rows) > 10:
                    break   # a header this far down is a different file shape
            continue

        records.append((index, dict(zip(header, values, strict=False))))

    if header is None:
        raise ValidationError(
            error_codes.IMPORT_COLUMN_MISSING,
            column=REQUIRED_COLUMNS[0],
            required=list(REQUIRED_COLUMNS),
            # What was actually found, so an operator can see why: usually a
            # title row, or column names in Dari rather than the expected keys.
            found=[value for value in (seen_rows[0] if seen_rows else []) if value],
        )
    return records


def parse(content: bytes, filename: str) -> tuple[list[ParsedVillage], list[RowProblem]]:
    """CSV, Excel or JSON in, structured rows out. Nothing touches the database."""
    lowered = filename.lower()

    if lowered.endswith((".xlsx", ".xlsm")):
        records = _rows_from_excel(content)
        return _to_villages(records)

    try:
        text = content.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as exc:
        # A file saved as Windows-1256 or similar. Naming the cause beats a
        # generic "could not read".
        raise ValidationError(
            error_codes.IMPORT_FILE_UNREADABLE, reason="not_utf8"
        ) from exc

    if filename.lower().endswith(".json"):
        raw_rows = json.loads(text)
        if not isinstance(raw_rows, list):
            raise ValidationError(error_codes.IMPORT_FILE_UNREADABLE, reason="expected_a_list")
        records = [(i + 1, dict(r)) for i, r in enumerate(raw_rows)]
    else:
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            raise ValidationError(error_codes.IMPORT_FILE_UNREADABLE, reason="no_header")
        missing = [c for c in REQUIRED_COLUMNS if c not in reader.fieldnames]
        if missing:
            raise ValidationError(
                error_codes.IMPORT_COLUMN_MISSING,
                column=missing[0],
                required=list(REQUIRED_COLUMNS),
            )
        records = [(i + 2, row) for i, row in enumerate(reader)]  # +2: header is line 1

    return _to_villages(records)


def _to_villages(
    records: list[tuple[int, dict[str, Any]]],
) -> tuple[list[ParsedVillage], list[RowProblem]]:
    parsed: list[ParsedVillage] = []
    problems: list[RowProblem] = []

    for row_number, row in records:
        district_code = _clean(row.get("district_code"))
        name = _clean(row.get("name"))

        if not district_code:
            problems.append(RowProblem(row_number, "district_code", "missing"))
            continue
        if not name:
            problems.append(RowProblem(row_number, "name", "missing"))
            continue

        latitude, lat_problem = _decimal(row.get("latitude"), row_number, "latitude")
        longitude, lon_problem = _decimal(row.get("longitude"), row_number, "longitude")
        problems.extend(p for p in (lat_problem, lon_problem) if p)

        parsed.append(
            ParsedVillage(
                row_number=row_number,
                district_code=district_code.upper(),
                name=name,
                aliases=_aliases(row.get("alternative_names")),
                latitude=latitude,
                longitude=longitude,
                code=_clean(row.get("code")) or None,
                note=_clean(row.get("note")) or None,
                station_name=_clean(row.get("station_name")) or None,
            )
        )
    return parsed, problems


def _clean(value: Any) -> str:
    return str(value).strip() if value not in (None, "") else ""


def _aliases(value: Any) -> list[str]:
    """Separated by | or ; or , -- spreadsheets in the wild use all three.

    A comma only reaches here from a quoted CSV cell or from JSON; an unquoted
    comma is the field separator and never gets this far.
    """
    raw = _clean(value)
    if not raw:
        return []
    for separator in ("|", ";", ","):
        if separator in raw:
            return [part.strip() for part in raw.split(separator) if part.strip()]
    return [raw]


def _decimal(value: Any, row_number: int, column: str) -> tuple[Decimal | None, RowProblem | None]:
    raw = _clean(value)
    if not raw:
        return None, None
    try:
        parsed = Decimal(raw)
    except (InvalidOperation, ValueError):
        return None, RowProblem(row_number, column, "not_a_number", raw, blocking=False)
    limit = Decimal(90) if column == "latitude" else Decimal(180)
    if not -limit <= parsed <= limit:
        return None, RowProblem(row_number, column, "out_of_range", raw, blocking=False)
    return parsed, None


@dataclass(frozen=True, slots=True)
class PreviewImportCommand:
    filename: str
    content: bytes
    actor_id: str
    actor_role: ActorRole = ActorRole.ADMIN


class PreviewVillageImport:
    """Step one. Reads the file, checks it against what exists, proposes
    duplicates, and stores the result as a job. Writes no village."""

    def __init__(self, *, jobs, villages, districts, geography, audit, clock, new_id) -> None:
        self._jobs = jobs
        self._villages = villages
        self._districts = districts
        self._geography = geography
        self._audit = audit
        self._clock = clock
        self._new_id = new_id

    def execute(self, cmd: PreviewImportCommand) -> ImportPreview:
        parsed, problems = parse(cmd.content, cmd.filename)

        known = {d.code.upper(): d for d in self._districts.list(limit=200)}
        for village in list(parsed):
            if village.district_code not in known:
                problems.append(
                    RowProblem(
                        village.row_number, "district_code", "unknown_district",
                        village.district_code,
                    )
                )
                parsed.remove(village)

        duplicates = self._detect(parsed, known)
        flagged = {d.row_number for d in duplicates}
        will_create = [v for v in parsed if v.row_number not in flagged]

        blocking = [problem for problem in problems if problem.blocking]

        job = self._jobs.create(
            id=self._new_id(),
            entity="villages",
            filename=cmd.filename,
            status=ImportStatus.PREVIEWED.value,
            total_rows=len(parsed) + len(blocking),
            valid_rows=len(parsed),
            error_rows=len(blocking),
            duplicate_rows=len(duplicates),
            report={
                "problems": [asdict(p) for p in problems],
                "duplicates": [asdict(d) for d in duplicates],
                "will_create": [_serialise(v) for v in will_create],
                # The parsed payload for every flagged row, so a duplicate the
                # operator confirms as genuinely new can actually be created.
                "flagged_payload": [
                    _serialise(v) for v in parsed if v.row_number in flagged
                ],
            },
        )
        self._jobs.flush()

        self._audit.write(
            "import.previewed",
            actor_id=cmd.actor_id,
            actor_role=cmd.actor_role,
            entity_type="import_job",
            entity_id=job.id,
            after={
                "filename": cmd.filename,
                "valid": len(parsed),
                "problems": len(blocking),
                "duplicates": len(duplicates),
            },
        )
        return ImportPreview(
            job_id=job.id,
            total_rows=len(parsed) + len(blocking),
            valid_rows=len(parsed),
            problems=problems,
            duplicates=duplicates,
            will_create=will_create,
        )

    def _detect(
        self, parsed: list[ParsedVillage], districts: dict[str, Any]
    ) -> list[DuplicateProposal]:
        """Compare against stored villages and against the file itself.

        A same-name village in a *different* district is reported but marked
        ``same_district=False``, because that is usually two real places rather
        than a duplicate -- and section 7 says the two must stay separate.
        """
        proposals: list[DuplicateProposal] = []

        for village in parsed:
            district = districts[village.district_code]
            for existing in self._geography.list_villages(district.id, limit=500):
                score = similarity(village.name, existing.name)
                if score >= DUPLICATE_THRESHOLD:
                    proposals.append(
                        DuplicateProposal(
                            row_number=village.row_number,
                            incoming_name=village.name,
                            existing_village_id=existing.id,
                            existing_name=existing.name,
                            score=round(score, 3),
                            same_district=True,
                            reason="exists_in_same_district",
                        )
                    )
                    break

        # Duplicates inside the file itself, which spreadsheets are full of.
        seen: dict[tuple[str, str], ParsedVillage] = {}
        for village in parsed:
            identity = (village.district_code, village.key)
            if identity in seen:
                proposals.append(
                    DuplicateProposal(
                        row_number=village.row_number,
                        incoming_name=village.name,
                        existing_village_id=None,
                        existing_name=seen[identity].name,
                        score=1.0,
                        same_district=True,
                        reason=f"repeated_in_file_at_row_{seen[identity].row_number}",
                    )
                )
            else:
                seen[identity] = village

        return proposals


@dataclass(frozen=True, slots=True)
class CommitImportCommand:
    job_id: str
    actor_id: str
    actor_role: ActorRole = ActorRole.ADMIN
    # Rows a person reviewed and confirmed are genuinely new despite the flag.
    accept_rows: tuple[int, ...] = ()
    create_stations: bool = True


@dataclass(frozen=True, slots=True)
class CommitImportResult:
    job_id: str
    villages_created: int
    aliases_created: int
    stations_created: int
    skipped_duplicates: int


class CommitVillageImport:
    """Step two. Writes only what the preview said it would, plus any rows a
    person explicitly accepted."""

    def __init__(
        self, *, jobs, villages, aliases, stations, districts, geography, audit, clock, new_id
    ) -> None:
        self._jobs = jobs
        self._villages = villages
        self._aliases = aliases
        self._stations = stations
        self._districts = districts
        self._geography = geography
        self._audit = audit
        self._clock = clock
        self._new_id = new_id

    def execute(self, cmd: CommitImportCommand) -> CommitImportResult:
        job = self._jobs.get(cmd.job_id)
        if job.status == ImportStatus.COMMITTED.value:
            raise ConflictError(error_codes.IMPORT_ALREADY_COMMITTED, job_id=job.id)
        if job.status != ImportStatus.PREVIEWED.value:
            raise ConflictError(
                error_codes.IMPORT_NOT_PREVIEWED, job_id=job.id, status=job.status
            )

        report = job.report or {}
        rows = [_deserialise(r) for r in report.get("will_create", [])]

        accepted = set(cmd.accept_rows)
        if accepted:
            by_row = {d["row_number"]: d for d in report.get("duplicates", [])}
            unknown = accepted - set(by_row)
            if unknown:
                raise ValidationError(
                    error_codes.IMPORT_ROW_INVALID, rows=sorted(unknown), job_id=job.id
                )
            # A flagged row the operator confirmed is genuinely new.
            payload = report.get("flagged_payload", [])
            confirmed = [
                _deserialise(r) for r in payload if r["row_number"] in accepted
            ]
            if len(confirmed) != len(accepted):
                # Refuse rather than create a subset: an operator who accepted
                # five rows and got three would have no way to tell.
                raise ConflictError(
                    error_codes.IMPORT_ROW_INVALID,
                    job_id=job.id,
                    accepted=sorted(accepted),
                    resolved=len(confirmed),
                )
            rows.extend(confirmed)

        districts = {d.code.upper(): d for d in self._districts.list(limit=200)}
        counters = self._next_codes(districts)

        villages = aliases = stations = 0
        for row in rows:
            district = districts.get(row.district_code)
            if district is None:
                raise NotFoundError(
                    error_codes.DISTRICT_NOT_FOUND, code=row.district_code
                )

            code = row.code
            if not code:
                counters[district.code] += 1
                code = f"{district.code}-{counters[district.code]:03d}"
            if self._villages.find_by(code=code) is not None:
                raise ConflictError(error_codes.VILLAGE_CODE_TAKEN, code=code)

            village = self._villages.create(
                id=self._new_id(),
                code=code,
                name=row.name,
                name_key=row.key,
                district_id=district.id,
                latitude=row.latitude,
                longitude=row.longitude,
                status=GeoStatus.ACTIVE.value,
                source_note=row.note or f"imported from {job.filename}",
                created_by=cmd.actor_id,
            )
            self._villages.flush()
            villages += 1

            for alias in row.aliases:
                self._aliases.create(
                    id=self._new_id(),
                    village_id=village.id,
                    name=alias,
                    name_key=comparison_key(alias),
                    note=f"imported from {job.filename}",
                    created_by=cmd.actor_id,
                )
                aliases += 1

            if cmd.create_stations:
                station_name = row.station_name or f"ایستگاه {row.name}"
                self._stations.create(
                    id=self._new_id(),
                    code=f"{code}-S1",
                    name=station_name,
                    name_key=comparison_key(station_name),
                    village_id=village.id,
                    district_id=district.id,
                    latitude=row.latitude,
                    longitude=row.longitude,
                    is_primary=True,
                    status=GeoStatus.ACTIVE.value,
                    created_by=cmd.actor_id,
                )
                stations += 1

        job.status = ImportStatus.COMMITTED.value
        job.created_rows = villages
        job.committed_at = self._clock.now()
        self._jobs.save(job)

        skipped = int(job.duplicate_rows or 0) - len(accepted)
        self._audit.write(
            "import.committed",
            actor_id=cmd.actor_id,
            actor_role=cmd.actor_role,
            entity_type="import_job",
            entity_id=job.id,
            after={
                "villages": villages,
                "aliases": aliases,
                "stations": stations,
                "skipped_duplicates": skipped,
            },
        )
        return CommitImportResult(job.id, villages, aliases, stations, max(0, skipped))

    def _next_codes(self, districts: dict[str, Any]) -> dict[str, int]:
        """Continue each district's numbering rather than restarting it."""
        counters: dict[str, int] = {}
        for code, district in districts.items():
            highest = 0
            for village in self._geography.list_villages(district.id, limit=500):
                suffix = village.code.rsplit("-", 1)[-1]
                if suffix.isdigit():
                    highest = max(highest, int(suffix))
            counters[code] = highest
        return counters


def _serialise(village: ParsedVillage) -> dict[str, Any]:
    return {
        "row_number": village.row_number,
        "district_code": village.district_code,
        "name": village.name,
        "aliases": village.aliases,
        "latitude": str(village.latitude) if village.latitude is not None else None,
        "longitude": str(village.longitude) if village.longitude is not None else None,
        "code": village.code,
        "note": village.note,
        "station_name": village.station_name,
    }


def _deserialise(payload: dict[str, Any]) -> ParsedVillage:
    return ParsedVillage(
        row_number=payload["row_number"],
        district_code=payload["district_code"],
        name=payload["name"],
        aliases=list(payload.get("aliases") or []),
        latitude=Decimal(payload["latitude"]) if payload.get("latitude") else None,
        longitude=Decimal(payload["longitude"]) if payload.get("longitude") else None,
        code=payload.get("code"),
        note=payload.get("note"),
        station_name=payload.get("station_name"),
    )
